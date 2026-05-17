import asyncio
import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional

from app.config import supabase
from app.middleware.auth import admin_only

router = APIRouter()


def _build_excel(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Achievement Report"

    headers = [
        "Employee Name", "Department", "Manager",
        "Goal Title", "Thrust Area", "UoM Type", "Weightage (%)", "Target",
        "Q1 Actual", "Q1 Score (%)", "Q1 Status",
        "Q2 Actual", "Q2 Score (%)", "Q2 Status",
        "Q3 Actual", "Q3 Score (%)", "Q3 Status",
        "Q4 Actual", "Q4 Score (%)", "Q4 Status",
        "Weighted Score (%)",
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    row_num = 2
    for record in data["rows"]:
        row_fill = alt_fill if row_num % 2 == 0 else None
        values = [
            record.get("employee_name", ""),
            record.get("department", ""),
            record.get("manager_name", ""),
            record.get("goal_title", ""),
            record.get("thrust_area", ""),
            record.get("uom_type", ""),
            record.get("weightage", ""),
            record.get("target", ""),
            record.get("q1_actual", ""), record.get("q1_score", ""), record.get("q1_status", ""),
            record.get("q2_actual", ""), record.get("q2_score", ""), record.get("q2_status", ""),
            record.get("q3_actual", ""), record.get("q3_score", ""), record.get("q3_status", ""),
            record.get("q4_actual", ""), record.get("q4_score", ""), record.get("q4_status", ""),
            record.get("weighted_score", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = center_align
            if row_fill:
                cell.fill = row_fill
        row_num += 1

    # Auto-fit column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=0)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_length + 3, 35)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _fetch_report_data(cycle_id: str) -> dict:
    employees = supabase.table("profiles").select("id, full_name, department, manager_id")\
        .eq("role", "employee").order("full_name").execute()

    manager_ids = list({e["manager_id"] for e in employees.data if e.get("manager_id")})
    managers = {}
    if manager_ids:
        mgr_r = supabase.table("profiles").select("id, full_name").in_("id", manager_ids).execute()
        managers = {m["id"]: m["full_name"] for m in mgr_r.data}

    thrust_areas_r = supabase.table("thrust_areas").select("id, name").execute()
    thrust_areas = {t["id"]: t["name"] for t in thrust_areas_r.data}

    sheets = supabase.table("goal_sheets").select("id, employee_id, status")\
        .eq("cycle_id", cycle_id).execute()
    sheets_map = {s["employee_id"]: s for s in sheets.data}

    rows = []
    for emp in employees.data:
        sheet = sheets_map.get(emp["id"])
        if not sheet:
            rows.append({
                "employee_name": emp["full_name"],
                "department": emp.get("department", ""),
                "manager_name": managers.get(emp.get("manager_id", ""), ""),
                "goal_title": "(No goal sheet)",
            })
            continue

        goals_r = supabase.table("goals").select("*")\
            .eq("goal_sheet_id", sheet["id"]).order("sort_order").execute()

        if not goals_r.data:
            rows.append({
                "employee_name": emp["full_name"],
                "department": emp.get("department", ""),
                "manager_name": managers.get(emp.get("manager_id", ""), ""),
                "goal_title": "(No goals)",
            })
            continue

        for g in goals_r.data:
            ach_r = supabase.table("achievements").select("*")\
                .eq("goal_id", g["id"]).execute()
            ach_by_q = {a["quarter"]: a for a in ach_r.data}

            row = {
                "employee_name": emp["full_name"],
                "department": emp.get("department", ""),
                "manager_name": managers.get(emp.get("manager_id", ""), ""),
                "goal_title": g["title"],
                "thrust_area": thrust_areas.get(g.get("thrust_area_id", ""), ""),
                "uom_type": g["uom_type"],
                "weightage": g["weightage"],
                "target": g.get("target_value") or g.get("target_date") or "",
            }
            for q in ["Q1", "Q2", "Q3", "Q4"]:
                ach = ach_by_q.get(q)
                row[f"{q.lower()}_actual"] = ach["actual_value"] if ach and ach.get("actual_value") is not None else (ach["actual_date"] if ach else "")
                row[f"{q.lower()}_score"] = round(float(ach["computed_score"]), 1) if ach and ach.get("computed_score") is not None else ""
                row[f"{q.lower()}_status"] = ach["status"] if ach else ""

            scores = [float(ach_by_q[q]["computed_score"]) for q in ["Q1", "Q2", "Q3", "Q4"]
                      if q in ach_by_q and ach_by_q[q].get("computed_score") is not None]
            row["weighted_score"] = round(sum(scores) / len(scores), 1) if scores else ""
            rows.append(row)

    return {"rows": rows}


@router.get("/achievement-export")
async def export_achievement_report(
    cycle_id: Optional[str] = None,
    current_user: dict = Depends(admin_only),
):
    if not cycle_id:
        cycle_r = supabase.table("cycles").select("id").eq("is_active", True).maybe_single().execute()
        if not cycle_r.data:
            raise HTTPException(400, "No active cycle found")
        cycle_id = cycle_r.data["id"]

    loop = asyncio.get_event_loop()
    data = await loop.run_in_executor(None, _fetch_report_data, cycle_id)
    excel_bytes = await loop.run_in_executor(None, _build_excel, data)

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=achievement_report_{cycle_id}.xlsx"},
    )
